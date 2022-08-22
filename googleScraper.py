import random
import time

import openpyxl as op
import pandas as pd
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager

xl_name = "scraped1104.xlsx"
sheet_name = "URLs"

options = webdriver.ChromeOptions()
options.add_argument("--enable-javascript")
# useragent = UserAgent()
# userAgent = useragent.random
# options.add_argument(f'user-agent={userAgent}')
# options.add_argument("--headless")  # operate without visible instances of chrome
driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)

# the header used when retrieving text from the links
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.114 Safari/537.36"}

# Queries to obtain links
queries = ["free bitcoin generator",
           "bitcoin generator online",
           "free btc",
           "bitcoin doubler",
           "cryptocurrency generator",
           "free bitcoin hack",
           ]

links = []  # Empty list to capture results of crawl
scrapedtitles = []  # Contains either the name of the text file corresponding to the URL or a 'Connection refused' string, denoting that the site could not be crawled


# Specify number of pages on google search, each page contains 10 links
def googleScrape():
    # Iterate through list of queries, fetch links from random page and the first page given from search query.
    for index, query in enumerate(queries):
        n_pages = 7
        for page in range(1, n_pages):
            url = "http://www.google.com/search?q=" + queries[index] + "&start=" + str(
                (random.randint(2, 12) - 1) * 10)
            driver.get(url)
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            search = soup.find_all('div', class_="yuRUbf")
            # print(search)
            for h in search:
                links.append(h.a.get(
                    'href'))

        time.sleep(
            3)

    # Iterate through list of queries, fetch results from first two pages of each query at 10 links per page.
    for index, query in enumerate(queries):
        n_pages = 7
        for page in range(1, n_pages):

            url = "http://www.google.com/search?q=" + queries[index] + "&start=" + str((page - 1) * 10)
            driver.get(url)
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            search = soup.find_all('div', class_="yuRUbf")
            for h in search:
                links.append(h.a.get('href'))
            time.sleep(
                3)
            print(links)
        time.sleep(
            3)


# iterates through list of links, retrieves text displayed to the user
def scrapeText(links):
    for index, url in enumerate(links):
        # url = "http://" + url #for lists without protocol appended, only used when scraping bitcoin.fr
        printOutput = []
        try:
            url = requests.get(
                url, headers=headers)  # makes a request to a url in the list
            url_content = url.content

            soup_url = BeautifulSoup(url_content,
                                     'html.parser')  # initialise beautifulsoup, it is later called via soup_url
            text = soup_url.get_text(" ",
                                     strip=True)  # extracts text from page, does not work 100% of the time but has a higher success rate than other methods
            with open("file%d.txt" % index, "w",
                      encoding="utf-8") as f:  # creates new text file with text and index from url
                f.write(str(text))  # + "\n",
            scrapedtitles.append("file%d.txt" % index)
            printOutput.append(text)
            print(scrapedtitles)
            # print(printOutput) #used to print the resulting text into the console

            time.sleep(
                3)  # allow time for page to load and to prevent bot detection, adds 3 seconds per link

        except requests.exceptions.ConnectionError:  # skips the url if the connection fails
            requests.status_code = "Connection refused"
            scrapedtitles.append("Connection refused")
            # index+1
            continue


# Load the list into excel file using pandas
def updateExcel(urllst, txtlst):
    tuple(urllst)
    tuple(txtlst)
    # Load list to dataframe
    df = pd.DataFrame(list(zip(urllst, txtlst)), columns=('URL', 'Text'))  #
    df.index += 1  # responsible for the 55/56 error as well as the int/str/int error

    # Write dataframe to excel
    xlw = pd.ExcelWriter(xl_name)
    df.to_excel(xlw, sheet_name=sheet_name, index_label="#", columns=["URL", "Text"])
    xlw.save()


def formatExcel(xl, sheet="Sheet1"):
    # open the excel file
    workBook = op.load_workbook(xl)
    workSheet = workBook.get_sheet_by_name(sheet)

    # set column width
    cols = ("A", "B", "C")
    widths = (5, 80, 25)

    # colour formatting
    colourFill = op.styles.PatternFill(start_color="5F9F9F",
                                      fill_type='solid')

    # define border style
    thin_border = op.styles.borders.Border(left=op.styles.Side(style='thin'),
                                           right=op.styles.Side(style='thin'),
                                           top=op.styles.Side(style='thin'),
                                           bottom=op.styles.Side(style='thin'))

    # define Text wrap
    text_wrap = op.styles.Alignment(wrap_text=True)

    # Format the header row
    for row in range(1, 2):  # Loop only the 1st row
        for col in range(1, workSheet.max_column + 1):  # loop through all columns
            workSheet.cell(row=row, column=col).fill = colourFill

    # format cells
    for row in workSheet.iter_rows():
        for cell in row:
            # apply borders
            cell.border = thin_border
            # apply wrap
            cell.alignment = text_wrap

    # save file
    workBook.save(xl)


googleScrape()
print(links)
print("Number of links in the list: ", len(links), "\n")
links = list(dict.fromkeys(links))  # delete duplicates
number_of_elements = len(links)
print("Number of links in the list without duplicates: ", number_of_elements)
print(scrapeText(links))

# Load the match list to excel
updateExcel(links, scrapedtitles)

# Format the excel file
formatExcel(xl_name, sheet_name)
